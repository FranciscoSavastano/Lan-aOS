from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry
from datetime import date
from xlwt import *
import openpyxl

root = Tk()

class app(): 
    def __init__(self):
        self.root = root
        self.window()
        self.subwindow()
        self.botoes()
        root.mainloop()        
        
    
    def window(self): #cria a janela principal
        self.root.title('Lança OS')
        self.root.configure(background='#ffffff')
        self.root.geometry("600x700")
        self.root.resizable(False, False)
        
    def subwindow(self): #cria as subjanelas para os widgets, e os insere
        self.frame_1 = Frame()
        self.frame_1 = Label(
            bg='#cac3ba', bd=2, highlightbackground='#000000', highlightthickness=1)
        self.frame_1.place(relx=0.04, rely=0.05, relwidth=0.9, relheight=0.9)
        self.frame_botoes = Frame()
        self.frame_botoes.place(relx=0.04, rely= 0.05, relwidth=0.9, relheight=0.9)
        entrytext = Label(
            text='Codigo do produto: \n\n\nNumero da Nota: \n\n\nData de emissão: \n\n\nNumero da OS: \n\n\nTipo de operação: \n\n\n\n\nProblema:\n\n\n\n\nNumero da NS:\n\n\n\n Consultor(a):', font=('verdana', 11))
        entrytext.place(relx=0.06, rely=0.134)
    def botoes(self):
        global entryCodPar, entryOs, entryNF, entryEmissao, entryNS, entryConsul, entryTroca, entryProblema
        entryCodPar = Entry(self.frame_botoes)
        entryCodPar.place(relx= 0.4, rely= 0.1)
        entryNF = Entry(self.frame_botoes)
        entryNF.place(relx=0.4, rely = 0.185)
        entryEmissao = DateEntry(self.frame_botoes, locale = "pt_BR")
        entryEmissao.place(relx= 0.4, rely= 0.27)        
        entryOs = Entry(self.frame_botoes)
        entryOs.place(relx= 0.4, rely= 0.36)
        entryTroca = ttk.Combobox(        state="readonly",
        values=["Troca defeito", "Em Teste", "Compra Errada", "Outros"]
        )
        entryTroca.place(relx= 0.4, rely= 0.45)
        entryProblema = Text(self.frame_botoes, height= 4, width= 35)
        entryProblema.place(relx= 0.4, rely= 0.5)
        entryNS = Text(self.frame_botoes, height= 4, width= 35)
        entryNS.place(relx= 0.4, rely= 0.65)
        entryConsul = Entry(self.frame_botoes)
        entryConsul.place(relx= 0.4, rely= 0.84)
        enviar = Button(text = ("Enviar"), command= self.Oswin)
        enviar.place(relx= 0.4, rely= 0.9)
    def Oswin(self):
        global CodPar, Nf, Emissao, Os, Troca, Problema, Ns, Consultor
        CodPar = entryCodPar.get()
        Nf = entryNF.get()
        Emissao = entryEmissao.get()
        Os = entryOs.get()
        Troca = entryTroca.get()
        Problema = entryProblema.get("1.0", END)
        Ns = entryNS.get("1.0", END)
        Consultor = entryConsul.get()
        win = Toplevel()
        win.geometry('600x600')
        win.resizable(False, False)
        res_1 = Frame(win)
        res_1.place(relx=0, rely=0, relheight=1, relwidth=1)
        lab_1 = Label(res_1, bg='#d3d3d3',
                      text = f"Codigo do produto : {CodPar}\n Numero da Nota: {Nf}\n Emissão da Nota: {Emissao}\n Numero da OS: {Os} \n Tipo de operação: {Troca}\n Problema: {Problema} \n NS do produto: {Ns}")
        lab_1.place(relheight=1, relwidth=1)
        res_2 = Frame(win, bg='#d3d3d3' )
        lab_2 = Label(res_2, bg='#d3d3d3', text = f"Operador \n\n\n Valor\n\n\n Unidade \n\n\n Nu \n\n\n Codigo do Cliente" ,font=('verdana', 9))
        lab_2.place(relx = 0.2, rely = 0.10)
        
        res_2.place(relx = 0, rely = 0, relheight=0.35, relwidth=1)
        global entryOperador, entryValor, entryUnidade, entryNu, entryCodCliente, entryChaveAcesso, entryNotaDev
        entryOperador = Entry(res_2)
        entryValor = Entry(res_2)
        entryUnidade = ttk.Combobox(res_2,       state="readonly",
        values=["CENTRO", "BONSUCESSO", "VILAR", "TAQUARA", "CAMPO GRANDE", "CABO FRIO"]
        )
        entryNu = Entry(res_2)
        entryCodCliente = Entry(res_2)
        if(Troca == "Compra Errada"):
            lab_3 = Label(res_2, bg="#d3d3d3", text = "Chave de Acesso\n\n\n Num. Nota de devolução", font=('verdana', 9))
            lab_3.place(relx= 0.7, rely = 0.10)
            entryChaveAcesso = Entry(res_2)
            entryNotaDev = Entry(res_2)
            entryChaveAcesso.place(relx = 0.5, rely = 0.10)
            entryNotaDev.place(relx = 0.5, rely = 0.294)
        else:
            entryChaveAcesso = None
            entryNotaDev = None
        entryOperador.place(relx = 0, rely = 0.10)
        entryValor.place(relx = 0, rely = 0.29)
        entryUnidade.place(relx = 0, rely = 0.49)
        entryNu.place(relx = 0, rely = 0.68)
        entryCodCliente.place(relx = 0, rely = 0.894)
        writetxt = Button(res_1, text = ("Salvar como texto"), command= self.Writetxt)
        writetxt.place(relx= 0.2, rely= 0.9)
        writeexcel = Button(res_1,text = ("Salvar em excel"), command= self.Writeexc)
        writeexcel.place(relx= 0.5, rely= 0.9)
    def Writetxt(self):
        file = open(f"Os {Os} {Troca}.txt", "x")
        file.write(f"Numero da OS: {Os}\n")
        file.write(f"Numero da NF: {Nf}     Emissão: {Emissao}\n")
        file.write(f"Problema: {Problema}\n")
        file.write(f"Codigo do produto: {CodPar}\n")
        file.write(f"NS do produto: {Ns}\n")
        file.write(f"Cliente de {Consultor}\n")
        
        file.close()
    def Writeexc(self):
        CodCliente =Valor =Unidade =Nu =Operador = ChaveAcesso = NumDev = None
        CodCliente = entryCodCliente.get()
        Valor = entryValor.get()
        Unidade = entryUnidade.get()
        Nu = entryNu.get()
        Operador = entryOperador.get()
        if(entryChaveAcesso and entryNotaDev != None ): 
            
            ChaveAcesso = entryChaveAcesso.get()
            NumDev = entryNotaDev.get()
        
        try:
            file = "C:/Users/MAKE-LAB/Documents/Controle Compartilhado os.xlsx"
            wb = openpyxl.load_workbook(filename=file)
            ws = wb["Planilha1"]
            lvazia = None
            cellval = None
            line = 4
            while lvazia == None:
                cellval = ws[f"B{line}"].value
                if cellval == None:
                    lvazia = line
                    x = 2
                    while x < 15:
                        coord = ws.cell(row = line , column= x)

                        if(coord.value == None):
                            pass
                        else:
                           line += 1
                           x = 2
                        x += 1
                    print("Não foi encontrado items")
                #else:
                line += 1
            campos = [Os, date.today(), Nf, NumDev, ChaveAcesso, CodCliente, Emissao, Nu, Troca, Valor, Consultor, Unidade, Operador, "-"]  
            for i in range(len(campos)):
                print(campos[i])
                if(campos[i] == None):
                    campos[i] = "-"
                elif (campos[i] == ""):
                    campos[i] = "-"
                    
            line -= 1
            x = 2
            for i in range(13):
                
                ws.cell(line, x).value = campos[i]
                x += 1
            print(line, x)
            print(ws.cell(line, x).value)
            
            wb.save("teste.xlsx")
        except Exception as e: print(e)
        # Workbook is created
            #wb = Workbook()
        
        # add_sheet is used to create sheet.
        
        
app()